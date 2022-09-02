import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
#
# serv_obj = Service("C:\\drivers\\chromedriver.exe")
# driver = webdriver.Chrome(service=serv_obj)
#
# driver.get("https://www.opencart.com")
# driver.maximize_window()

# file-workbook-sheet-rows and columns
# file = "C:\drivers\data.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet"]
#
# # counting no of rows in a file
# numRows = sheet.max_row
# # counting number of columns in a file ,
# numColumns = sheet.max_column
#
# # Reading all rows and coln from excel sheet,# Reading all rows and coln from excel sheet,
# for r in range(1, numRows + 1):
#     for c in range(1, numColumns + 1):
#         print(sheet.cell(r, c).value , end="     ")  # this syntax will get you the values from a cell. end will provide space between data
#     print()
#
# write data into a excel file
'''
file = "C:\\drivers\\writedatainthis.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active # or sheet = workbook["Sheet1"]   -- get active sheet from file

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value = "Welcome"

workbook.save(file) # never skip this line else all data will be lost.
'''
# mutiple data
file = "C:\drivers\mutipledatawrite.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active # or sheet = workbook["Sheet1"]   -- get active sheet from file


sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "Sanjay"
sheet.cell(1,3).value = "Engineer"

sheet.cell(2,1).value = 234
sheet.cell(2,2).value = "Sam"
sheet.cell(2,3).value = "doctor"

sheet.cell(3,1).value = 453
#sheet.cell(3,2).value = "Vrish"
sheet.cell(3,2).value = "kriiiisshupdate name"
sheet.cell(3,3).value = "Killer"

workbook.save(file)
