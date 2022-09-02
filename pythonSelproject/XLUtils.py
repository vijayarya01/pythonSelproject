import openpyxl
from openpyxl.styles import PatternFill

def readRow_Count(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumn_Count(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,coulmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,coulmnno).value

def writeData(file,sheetName,rownum,coulmnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,coulmnno).value = data
    workbook.save(file)


def fillGreenColor(file,sheetName,rownum,coulmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum,coulmnno).fill = greenFill
    workbook.save(file)


def fillRedColor(file,sheetName,rownum,coulmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum,coulmnno).fill = redFill
    workbook.save(file)